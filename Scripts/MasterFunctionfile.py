import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import random 
import csv
import os
import pandas as pd
import requests
from bs4 import BeautifulSoup
import MasterFunctionfile as mf
from concurrent.futures import ThreadPoolExecutor, as_completed
import urllib
from scrapfly import ScrapflyClient, ScrapeConfig, ScrapeApiResponse
from datetime import datetime

def find_lowest_price_store_with_scrapfly(product_url):
    api_key = 'scp-test-6fc24c20fe1f4ba0a171e7355e9ab34f'  # Replace with your actual Scrapfly API key
    additional_stores = ['Outdoor Gear Exchange','The Deep End Swim and Sportswear','Metro Swim Shop','Elsmore Swim Shop','Xtreme Swim','Sears - BHFO', 'Shop Premium Outlets', 'Walmart - BHFO, Inc.','APerfectDealer', 'Van Dyke and Bacon', 'TC Running Co','eBay',"Macy's",'Kenco Outfitters','Grivet Outdoors','TravelCountry.com','EMS','Famous Brands','Walmart - BuyBox Club','Baseball Savings.com','Sears - Ricci Berri', 'Slam Jam', 'mjfootwear.com', 'Sports Basement', 'ModeSens','Runnerinn.com','ShoeVillage.com','Running Zone','The Heel Shoe Fitters','Nikys Sports',"Beck's Shoes",'Next Step Athletics', "Brown's Shoe Fit Co. Dubuque", 'Super Shoes', 'JosephBeauty', 'Pants Store', 'Fingerhut', "Brown's Shoe Fit Co. Longview", 'Deporvillage.net' ]
    Black_List_Store = ["Chiappetta Shoes", "Sole Desire", "Shoe Station", "Bloomingdale's", "Lucky Shoes", "Glik's", "RushOrderTees", "Shoe Carnival", "FrontRunners LA", "Roderer Shoe Center", "Rogans Shoes", "Goodmiles Running Company", "Gazelle Sports", "Confluence Running", "Holabird Sports", "ssense.com", "Lyst", "Fleet Feet"]

    scrapfly = ScrapflyClient(key=api_key)
    
    error_counter = 0
    
    try:
        result: ScrapeApiResponse = scrapfly.scrape(ScrapeConfig(
            tags=["player", "project:default"],
            country="us",
            asp=True,
            render_js=True,
            rendering_wait=3000,
            url=product_url
        ))
        
        if result.status_code == 200:
            soup = BeautifulSoup(result.content, 'html.parser')
            lowest_price = None
            lowest_price_store = None

            offer_rows = soup.find_all('tr', class_='sh-osd__offer-row')

            for row in offer_rows:
                try:
                    store_name_tag = row.find('a', class_='b5ycib shntl')
                    quality_tag = row.find('span', class_='P6GC4b')
                    price_tag = row.find('span', class_='g9WBQb fObmGc')
                    shipping_info = row.find('td', class_='SH30Lb yGibJf')

                    if store_name_tag and price_tag and shipping_info:
                        store_name_raw = store_name_tag.get_text(strip=True)
                        store_name = store_name_raw.split('Opens')[0].strip()
                        
                        # Check if the store is blacklisted
                        if any(bad_word in store_name for bad_word in Black_List_Store):
                            continue  # Skip this store and move to the next one
                        
                        price_str = price_tag.get_text(strip=True).replace('$', '').replace(',', '').replace('\xa0€', '').strip()
                        price = float(price_str)

                        if 'amazon' in store_name.lower():
                            continue

                        shipping_text = shipping_info.get_text(strip=True).lower()
                        shipping_cost = calculate_shipping_cost(shipping_text)

                        total_price = price + shipping_cost
                        total_price = total_price

                        if is_preferred_store(quality_tag, store_name, additional_stores):
                            if lowest_price is None or total_price < lowest_price:
                                lowest_price = lowest_price = round(total_price, 2)
                                lowest_price_store = store_name
                except Exception as inner_exc:
                    error_counter += 1
                    print(f"Error processing row: {inner_exc}")
                    continue

            return lowest_price_store, lowest_price
        else:
            error_counter += 1
            print(f"Scrapfly request failed with status code {result.status_code}")
    except Exception as e:
        error_counter += 1
        print(f"An error occurred: {e}")

    return None, None

def log_to_file(message):
    with open('ErrorHandling.txt', 'a') as f:
        f.write(message + '\n')
    f.close()

def merge_csv(input_file):
    # Specify the output file name
    output_file = "Keepa_Combined_Export.csv"
    
    # Check if the output file exists
    if os.path.exists(output_file):
        # Load existing data from output file
        existing_data = pd.read_csv(output_file)
    else:
        # Create an empty DataFrame if the output file doesn't exist
        existing_data = pd.DataFrame()

    # Read data from the input file
    input_data = pd.read_csv(input_file)

    # If 'ASIN' column exists in the existing data, remove duplicates based on 'ASIN'
    if 'ASIN' in existing_data.columns:
        input_data = input_data[~input_data['ASIN'].isin(existing_data['ASIN'])]

    # Append new data to existing data
    merged_data = pd.concat([existing_data, input_data], ignore_index=True)

    # Write merged data to the output file
    merged_data.to_csv(output_file, index=False)

    print(f"Data from {input_file} merged successfully.")



def find_lowest_price_store_with_scrapingbee(product_url):
    api_key = '7WX8OW6NE4303BOGNO05RWHKP0COKSO0ZXZDF4Y0FKBZ5G7HDS5276Z1MCV9IU2EFRQC14OCU4AR7VI7'
    additional_stores = ['Sears - BHFO', 'Shop Premium Outlets', 'Walmart - BHFO, Inc.','APerfectDealer', 'Van Dyke and Bacon', 'TC Running Co','eBay',"Macy's",'Kenco Outfitters','Grivet Outdoors','TravelCountry.com','EMS','Famous Brands','Walmart - BuyBox Club','Baseball Savings.com']
    try:
        response = requests.get(
            'https://app.scrapingbee.com/api/v1/',
            params={
                'api_key': api_key,
                'url': product_url,
                'render_js': 'true',
                'custom_google': 'true',
                'wait_for': 'a[class*=title]'
            }
        )
    
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            lowest_price = None
            lowest_price_store = None

            offer_rows = soup.find_all('tr', class_='sh-osd__offer-row')

            for row in offer_rows:
                try:
                    store_name_tag = row.find('a', class_='b5ycib shntl')
                    quality_tag = row.find('span', class_='P6GC4b')
                    price_tag = row.find('span', class_='g9WBQb fObmGc')
                    shipping_info = row.find('td', class_='SH30Lb yGibJf')

                    if store_name_tag and price_tag and shipping_info:
                        store_name_raw = store_name_tag.get_text(strip=True)
                        store_name = store_name_raw.split('Opens')[0].strip()
                        # Adapted to handle different currency formats and symbols
                        price_str = price_tag.get_text(strip=True).replace('$', '').replace(',', '').replace('\xa0€', '').strip()
                        price = float(price_str)

                        if 'amazon' in store_name.lower():
                            continue

                        shipping_text = shipping_info.get_text(strip=True).lower()
                        shipping_cost = calculate_shipping_cost(shipping_text)

                        total_price = price + shipping_cost
                        total_price = round(total_price, 2)

    
                        if is_preferred_store(quality_tag, store_name, additional_stores):
                            if lowest_price is None or total_price < lowest_price:
                                lowest_price = total_price
                                lowest_price_store = store_name
                except Exception as inner_exc:
                    # Handle any errors in parsing individual offer rows
                    continue

            return lowest_price_store, lowest_price
        else:
            print(f"ScrapingBee request failed with status code {response.status_code}")
            return None, None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None

def is_preferred_store(quality_tag, store_name, additional_stores):
    is_top_quality = quality_tag and quality_tag.get_text(strip=True) == 'Top Quality Store'
    is_additional_store = any(store in store_name for store in additional_stores)
    return is_top_quality or is_additional_store

def calculate_shipping_cost(shipping_text):
    if 'Spend' in shipping_text:
        return 10.00
    elif 'Free delivery' in shipping_text or 'Free shipping' in shipping_text:
        return 0
    elif '$' in shipping_text:
        cost = float(shipping_text.split('$')[1].split()[0])
        if cost > 20:
            return 10.00
        else:
            return cost
    else:
        return 0
    
def update_amazon_listing_price(input_file_path, output_file_path):
    # Load the DataFrame from the saved CSV
    df = pd.read_csv(input_file_path)

    # Iterate over each row in DataFrame
    for index, row in df.iterrows():
        if pd.notna(row['Min Price']):  # Check if 'Min Price' has a value
            amazon_list_price = calculate_amazon_list_price(row)
            if amazon_list_price is not None:
                # Update values if amazon_list_price is calculated
                df.at[index, 'Amazon_List_price'] = amazon_list_price
                df.at[index, 'Can_List'] = 'Y'
                df.at[index, 'Curr_Listed?'] = 2
            else:
                # Ensure that Amazon_List_price is explicitly set to None if not calculated
                df.at[index, 'Amazon_List_price'] = None

    # Save the updated DataFrame to CSV
    df.to_csv(output_file_path, index=False)
    print("Updated DataFrame saved")

def calculate_amazon_list_price(row):
    # Convert to float and handle None
    def to_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    buy_box_current = to_float(row.get('Buy Box: Current'))
    new_current = to_float(row.get('New: Current'))
    new_highest = to_float(row.get('New: Highest'))
    min_price = to_float(row.get('Min Price'))
    max_price = to_float(row.get('Max Price'))

    # Check for Lowest FBM Seller with specific ID
    #if row.get('Lowest FBM Seller') == 'ANZYNJW9IIF9C':
        #if new_current and new_current > min_price:
            #return new_current

    # Check Buy Box: Current if it's available and meets the minimum price requirement
    if buy_box_current and min_price and buy_box_current > min_price:
        return buy_box_current - 0.10

    # If Buy Box is not available, check New: Current if it's available and meets the minimum price requirement
    if new_current and min_price and new_current > min_price:
        return new_current - 0.10

    # If either number is available and not greater than Min Price then return none
    if (buy_box_current is not None or new_current is not None):
        if (buy_box_current is not None and buy_box_current < min_price) or \
           (new_current is not None and new_current < min_price):
            return None

    # If neither is available, check if Max Price < New: Highest
    if new_highest:
        if max_price < new_highest:
            return new_highest
        elif new_highest < min_price:
            return None
        elif min_price < new_highest < max_price:
            return new_highest

    # If there is no New: Highest price then return Max Price
    return max_price if max_price else None

def update_amazon_listing_price(input_file_path, output_file_path):
    # Load the DataFrame from the saved CSV
    df = pd.read_csv(input_file_path)

    # Iterate over each row in DataFrame
    for index, row in df.iterrows():
        if pd.notna(row['Min Price']):  # Check if 'Min Price' has a value
            # Attempt to calculate the Amazon list price
            amazon_list_price = calculate_amazon_list_price(row)
            if amazon_list_price is not None:
                # Update values if amazon_list_price is calculated
                df.at[index, 'Amazon_List_price'] = amazon_list_price
                df.at[index, 'Can_List'] = 'Y'
                df.at[index, 'Curr_Listed?'] = 2
            else:
                # Set values to reflect that a valid Amazon list price wasn't calculated
                df.at[index, 'Amazon_List_price'] = None
                df.at[index, 'Can_List'] = 'N'
                df.at[index, 'Curr_Listed?'] = 0
        else:
            # Set values for rows where 'Min Price' is not available
            df.at[index, 'Amazon_List_price'] = None
            df.at[index, 'Can_List'] = 'N'
            df.at[index, 'Curr_Listed?'] = 0

    # Save the updated DataFrame to CSV
    df.to_csv(output_file_path, index=False)
    print("Updated DataFrame saved")

def Update_Can_List(row):
    import pandas as pd  # Ensure pandas is imported

    # If 'Min Price' is NaN, return "N"
    if pd.isna(row['Min Price']):
        return "N"

    # Attempt to convert 'Buy Box: Current' and 'New: Current' to float, handling exceptions
    try:
        buy_box_current = float(row['Buy Box: Current']) if pd.notna(row['Buy Box: Current']) else None
    except ValueError:
        buy_box_current = None

    try:
        new_current = float(row['New: Current']) if pd.notna(row['New: Current']) else None
    except ValueError:
        new_current = None

    # Determine if 'Min Price' is greater than either 'Buy Box: Current' or 'New: Current'
    if buy_box_current is not None or new_current is not None:
        if (buy_box_current is not None and row['Min Price'] > buy_box_current) or \
           (new_current is not None and row['Min Price'] > new_current):
            return "N"
    return "Y"

def clean_and_convert_price(value):
    if pd.notna(value):
        if isinstance(value, str):
            cleaned_value = value.replace('$', '').replace(',', '').strip()
            try:
                return round(float(cleaned_value), 2)
            except ValueError:
                return None
        elif isinstance(value, float):
            return round(value, 2)
    return None


def Update_DB_After_ManualV(input_file, master_db_file, master_v_file, Black_Listed_Y):
    # Load master database
    master_db = pd.read_csv(master_db_file)

    # Load or initialize MasterV file
    if os.path.exists(master_v_file):
        master_v = pd.read_csv(master_v_file)
    else:
        master_v = pd.DataFrame(columns=master_db.columns)
        master_v['Curr_match'] = 0

    # Initialize or load Black_Listed_Y file
    if os.path.exists(Black_Listed_Y):
        black_listed_y = pd.read_csv(Black_Listed_Y)
    else:
        black_listed_y = pd.DataFrame(columns=master_db.columns)

    # Load input file
    input_df = pd.read_csv(input_file)

    # Initialize counters for items added to master_v_file and black_listed_y file
    items_added_to_master_v = 0
    items_added_to_black_listed_y = 0

    # Process each ASIN from the input file
    for asin in input_df['ASIN']:
        # Find if the ASIN exists in master_db and its Blacklisted status
        asin_row = master_db[master_db['ASIN'] == asin]

        # If ASIN is found in master_db
        if not asin_row.empty:
            # Check if any of the entries for this ASIN in the 'Blacklisted' column contain 'Y' or 'V'
            blacklisted_entries = asin_row['Blacklisted'].str.upper().str.strip()
            if 'Y' in blacklisted_entries.values:
                # If ASIN is blacklisted with 'Y' and not already in Black_Listed_Y
                if asin not in black_listed_y['ASIN'].values:
                    new_row = asin_row.copy()
                    black_listed_y = pd.concat([black_listed_y, new_row], ignore_index=True)
                    items_added_to_black_listed_y += 1
            elif 'V' in blacklisted_entries.values:
                # If ASIN is blacklisted with 'V' and not already in master_v_file
                if asin not in master_v['ASIN'].values:
                    new_row = asin_row.copy()
                    master_v = pd.concat([master_v, new_row], ignore_index=True)
                    items_added_to_master_v += 1

    # If ASIN is not found or not blacklisted with 'V' or 'Y', add to manual_check
    for asin in input_df['ASIN']:
        if asin not in black_listed_y['ASIN'].values and asin not in master_v['ASIN'].values:
            input_row = input_df[input_df['ASIN'] == asin].copy()
            black_listed_y = pd.concat([black_listed_y, input_row], ignore_index=True)
            items_added_to_black_listed_y += 1

    # Remove duplicates based on 'ASIN' in black_listed_y
    black_listed_y.drop_duplicates(subset=['ASIN'], keep='last', inplace=True)

    # Save the updated files
    black_listed_y.to_csv(Black_Listed_Y, index=False)
    master_v.to_csv(master_v_file, index=False)

    # Print the number of items added to master_v_file and black_listed_y file
    print(f"{items_added_to_master_v} items were added to master_v_file.")
    print(f"{items_added_to_black_listed_y} items were added to black_listed_y file.")

    print("Files have been updated.")
    
import pandas as pd

def create_manualcheck_file(input_file):
    # Load MasterV.csv
    master_df = pd.read_csv("DataBaseFiles\MasterV.csv")
    
    # Load input file
    input_df = pd.read_csv(input_file)
    
    # Initialize a list to store rows that need manual check
    manual_check_rows = []
    
    # Iterate through ASINs in the input file
    for index, row in input_df.iterrows():
        asin = row['ASIN']
        # Check if ASIN exists in MasterV.csv
        if asin not in master_df['ASIN'].values:
            manual_check_rows.append(row)
    
    # If there are rows needing manual check, create manualcheck.csv
    if manual_check_rows:
        manual_check_df = pd.DataFrame(manual_check_rows)
        manual_check_df.to_csv("ManualCheck.csv", index=False)
        print("Manual check file created: manualcheck.csv")
    else:
        print("All ASINs found in MasterV.csv. No manual check needed.")
# Example usage:
    #create_manualcheck_file("")
        
import pandas as pd

def update_master_db_w_manualcheck(master_db_path, update_csv_path):
    # Read Master_DB.csv
    master_db = pd.read_csv(master_db_path)

    # Read Update_CSV.csv
    update_csv = pd.read_csv(update_csv_path)

    # Iterate through rows in Update_CSV and update Master_DB accordingly
    for index, row in update_csv.iterrows():
        asin = row['ASIN']
        black_list = row.get('Black_list')
        product_code = row.get('PID')

        # Update Blacklist
        if black_list is not None:
            # Find the row in Master_DB that matches the ASIN
            master_row_index = master_db.index[master_db['ASIN'] == asin]

            # If ASIN found, update the Blacklisted column
            if not master_row_index.empty:
                master_db.loc[master_row_index, 'Blacklisted'] = black_list

        # Update Link
        if product_code is not None and pd.notna(product_code) and str(product_code).lower() != "none":
            # Find the row in Master_DB that matches the ASIN
            master_row_index = master_db.index[master_db['ASIN'] == asin]

            # If ASIN found, update the Extraction_Link column
            if not master_row_index.empty:
                pid = product_code
                upc = master_db.loc[master_row_index, 'Product Codes: UPC'].iloc[0]  # Extract UPC from Master_DB
                new_link = f'https://www.google.com/shopping/product/{pid}/offers?q={upc}&prds=cid:{pid},cond:1'
                master_db.loc[master_row_index, 'Extraction_Link'] = new_link

    # Save the updated Master_DB
    master_db.to_csv(master_db_path, index=False)

    print("Master_DB updated successfully.")



def update_BlackList(master_db_path, update_csv_path):
    # Read Master_DB.csv
    master_db = pd.read_csv(master_db_path)

    # Read Update_CSV.csv
    update_csv = pd.read_csv(update_csv_path)

    # Iterate through rows in Update_CSV and update Master_DB accordingly
    for index, row in update_csv.iterrows():
        asin = row['ASIN']
        black_list = row['Black_list']

        # Find the row in Master_DB that matches the ASIN
        master_row_index = master_db.index[master_db['ASIN'] == asin]

        # If ASIN found, update the Blacklisted column
        if not master_row_index.empty:
            master_db.loc[master_row_index, 'Blacklisted'] = black_list

    # Save the updated Master_DB
    master_db.to_csv(master_db_path, index=False)

def update_link(master_db_path, update_csv_path):
    # Read Master_DB.csv
    master_db = pd.read_csv(master_db_path)

    # Read Update_CSV.csv
    update_csv = pd.read_csv(update_csv_path)

    # Iterate through rows in Update_CSV and update Master_DB accordingly
    for index, row in update_csv.iterrows():
        asin = row['ASIN']
        product_code = row['PID']

        # Skip the ASIN if the product code is None or "None"
        if pd.isna(product_code) or str(product_code).lower() == "none":
            continue

        # Find the row in Master_DB that matches the ASIN
        master_row_index = master_db.index[master_db['ASIN'] == asin]
        print("Master row index for ASIN", asin, ":", master_row_index)

        # If ASIN found, update the Extraction_Link column
        if not master_row_index.empty:
            pid = product_code
            upc = master_db.loc[master_row_index, 'Product Codes: UPC'].iloc[0]  # Extract UPC from Master_DB
            new_link = f'https://www.google.com/shopping/product/{pid}/offers?q={upc}&prds=cid:{pid},cond:1'
            print("New link for ASIN", asin, ":", new_link)
            master_db.loc[master_row_index, 'Extraction_Link'] = new_link

    # Save the updated Master_DB
    master_db.to_csv(master_db_path, index=False)

import pandas as pd
from datetime import datetime
import os

def update_listing_stats(master_csv_path, results_csv_path):
    """
    Update and save the listing statistics based on the master CSV file.

    Args:
    master_csv_path (str): Path to the master CSV file.
    results_csv_path (str): Path to save the updated results CSV file.
    """
    # Read the data from the CSV file
    df = pd.read_csv(master_csv_path)

    # Group by the "Brand" column and count the number of rows for each brand
    brand_counts_all = df.groupby('Brand').size()

    # Filter the DataFrame where 'Curr_Listed?' equals 1
    filtered_df = df[df['Curr_Listed?'] == 2]

    # Group by the "Brand" column in the filtered DataFrame and count the number of rows for each brand
    brand_counts_curr_listed = filtered_df.groupby('Brand').size()

    # Get the current date
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Create a DataFrame to store the results
    results_df = pd.DataFrame({
        'Brand': brand_counts_all.index,
        'All_Listings_Count': brand_counts_all.values,
        'Current_Listings_Count': brand_counts_curr_listed.reindex(brand_counts_all.index, fill_value=0).values,
        'Date': [current_date] * len(brand_counts_all)
    })

    # Check if the result file exists to append or create new
    if os.path.exists(results_csv_path):
        # Read existing data
        existing_data = pd.read_csv(results_csv_path)
        
        # Combine existing data with new results
        combined_data = pd.concat([existing_data, results_df])
        
        # Drop duplicates: adjust subset to your needs for identifying duplicates
        clean_data = combined_data.drop_duplicates(subset=['Brand', 'Date'], keep='last')
        
        # Save the clean data back to the CSV file
        clean_data.to_csv(results_csv_path, index=False)
    else:
        # Create a new file and write headers if file does not exist
        results_df.to_csv(results_csv_path, index=False)

    print(f"\nUpdated results saved to '{results_csv_path}'")

# Example usage
update_listing_stats('DataBaseFiles/MasterV.csv', 'Results/Daily_listings_stats.csv')





def process_row_with_scrapingbee(row, index, brand_to_list):
    MIN_ROI = 1.7
    MAX_ROI = 1.87
    print(f"Starting processing of row {index} with asin {row['ASIN']}")
    log_to_file(f"Starting processing of row {index}")

    # Initialize default values for the updates
    update_info = {
        'Store': None,
        'Price': None,
        'Min Price': None,
        'Max Price': None,
        'Amazon_List_price': None,
        'Can_List': "N",  # Default to "N", will update if conditions are met
        'Curr_Listed?': 0
    }

    product_url = row['Extraction_Link']
    brand = row['Brand'].strip().lower() if pd.notna(row['Brand']) else ''
    brands_to_process = [brand.lower() for brand in brand_to_list]

    # Check if brand is eligible
    if brand in brands_to_process:
        log_to_file(f"Brand '{brand}' is eligible for processing.")
        # Now check if URL is valid
        if pd.notna(product_url):
            log_to_file(f"Brand: {brand} and ASIN: {row['ASIN']} passed the tests")
            print(f"Brand: {brand} and ASIN: {row['ASIN']} passed the tests")
            try:
                lowest_price_store, price = find_lowest_price_store_with_scrapfly(product_url)
                if price is not None:
                    min_price = round(price * MIN_ROI, 2)
                    max_price = round(price * MAX_ROI, 2)
                    update_info.update({
                        'Store': lowest_price_store,
                        'Price': price,
                        'Min Price': min_price,
                        'Max Price': max_price,
                    })

                    log_to_file(f"Row {index}: {row['ASIN']} Retrieved lowest price ${price} from store {lowest_price_store}. Calculated Min Price: ${min_price}, Max Price: ${max_price}")

                    calculated_value = calculate_amazon_list_price({**row.to_dict(), **update_info})
                    amazon_list_price = round(calculated_value, 2) if calculated_value is not None else None
                    log_to_file(f"Row {index}: Calculated Amazon List Price: ${amazon_list_price}")

                    if amazon_list_price is not None:
                        update_info.update({
                            'Amazon_List_price': amazon_list_price,
                            'Can_List': "Y",
                            'Curr_Listed?': 2
                        })
                    else:
                        log_to_file(f"ASIN {row['ASIN']}: No valid Amazon List Price found. Retaining default values.")
                else:
                    log_to_file(f"Row {index} {row['ASIN']}: No price found from scraping.")
            except Exception as e:
                log_to_file(f"Error processing row {index}: {e}")
        else:
            log_to_file(f"Row {index}: Invalid or missing URL for ASIN {row['ASIN']}")
    else:
        log_to_file(f"Row {index} (Brand: '{brand}') is not eligible for processing.")
    return (index, update_info)

def get_price_multiplier(price):
    if price <= 14.45:
        return 3
    elif price <= 29.45:
        return 2
    elif price <= 39.45:
        return 1.875
    elif price <= 49.45:
        return 1.84
    elif price <= 59.45:
        return 1.78
    elif price <= 79.45:
        return 1.7
    elif price <= 129.45:
        return 1.67
    elif price <= 159.45:
        return 1.50
    else:
        return 1.47

def keepa_asin_import(brands_2):
    # Reading the CSV file
    df = pd.read_csv('DataBaseFiles/MasterV.csv')
    
    # Filtering the DataFrame for the brands in brands_2
    filtered_df = df[df['Brand'].isin(brands_2)]
    
    # Writing the filtered DataFrame to a new CSV file, keeping the header
    filtered_df.to_csv('KeepaExports/Keepa_asin_import.csv', index=False)


def update_pricing_concurrently(Curr_Listed_path, master_db_path, Output_File_Price_Update, brand_to_list):
    Curr_Listed = pd.read_csv(Curr_Listed_path)
    master_db = pd.read_csv(master_db_path)  # Note: This is loaded but not used in this snippet.

    # Temporary storage for results that need additional processing
    update_list = []

    with ThreadPoolExecutor(max_workers=50) as executor:
        # Submit tasks for each row
        futures = [executor.submit(process_row_with_scrapingbee, row, index, brand_to_list) for index, row in Curr_Listed.iterrows()]

        # Collect all results before updating DataFrame
        for future in as_completed(futures):
            index, scrape_info = future.result()
            update_list.append((index, scrape_info))
    
    # Update DataFrame with scraped pricing info
    for index, scrape_info in update_list:
        for key, value in scrape_info.items():
            if key in ['Store', 'Price', 'Min Price', 'Max Price']:
                Curr_Listed.at[index, key] = value

    # Save the DataFrame with updated scrape data
    Curr_Listed.to_csv(Output_File_Price_Update, index=False)
    upload_file(r"C:\Users\Administrator\Documents\RA\DataBaseFiles\MasterV.csv")
    update_listing_stats('DataBaseFiles/MasterV.csv', 'Results/Daily_listings_stats.csv')
    print("Scrape complete and data saved")


def count_rows(csv_file):
    try:
        # Read the CSV file into a DataFrame
        df = pd.read_csv(csv_file)
        
        # Get the number of rows
        row_count = df.shape[0]
        
        print(f"Total rows in '{csv_file}': {row_count}")
    except FileNotFoundError:
        print(f"File '{csv_file}' not found.")



import pandas as pd
import os

def filter_and_export(csv_file, brand_names, output_filename):
    try:
        # Read the CSV file into a DataFrame
        df = pd.read_csv(csv_file)
        
        # Initialize an empty DataFrame to store filtered results
        filtered_df = pd.DataFrame()
        
        # Iterate over each brand name in the list
        for brand in brand_names:
            # Filter rows where Brand column matches the current brand name
            brand_df = df[df['Brand'].str.contains(brand)]
            # Append the filtered results to the overall filtered DataFrame
            filtered_df = pd.concat([filtered_df, brand_df])
        
        # Create the directory if it doesn't exist
        output_folder = "Filtered_outputs"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        
        # Construct the full output file path
        output_file_path = os.path.join(output_folder, output_filename)
        
        # Export the filtered DataFrame to a new CSV file
        filtered_df.to_csv(output_file_path, index=False)
        
        print(f"Filtered data exported to '{output_file_path}' successfully.")
    except FileNotFoundError:
        print(f"File '{csv_file}' not found.")


import pandas as pd

def update_extraction_links(csv_file):
    # Load CSV file into a pandas DataFrame
    df = pd.read_csv(csv_file)

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        pid = row['PID']
        if pid.lower() != 'none':  # Check if PID is not 'none'
            upc = row['Product Codes: UPC']
            new_link = f'https://www.google.com/shopping/product/{pid}/offers?q={upc}&prds=cid:{pid},cond:1'
            # Update the 'Extraction_Link' column with the new_link
            df.at[index, 'Extraction_Link'] = new_link

            # Append the row (excluding PID column) to MasterDB.csv
            row_without_pid = row.drop('PID')
            with open('DataBaseFiles\MasterV copy.csv', 'a') as f:
                f.write(','.join(map(str, row_without_pid)) + '\n')

            # Append the row (excluding PID column) to MasterV.csv
            with open('DataBaseFiles\Master_DB copy.csv', 'a') as f:
                f.write(','.join(map(str, row_without_pid)) + '\n')

    # Save the updated DataFrame back to the original CSV file
    df.to_csv(csv_file, index=False)
    print("done")


def update_blacklist_from_master_v(asin, value):
    # Define file paths
    master_v_path = "DataBaseFiles/MasterV.csv"
    master_db_path = "DataBaseFiles/Master_DB.csv"
    # Load MasterV.csv into a DataFrame
    master_v_df = pd.read_csv(master_v_path)
    # Check if ASIN exists in MasterV
    if asin not in master_v_df['ASIN'].values:
        print("Asin not found in MasterV")
        return
    # Remove ASIN column from MasterV
    master_v_df.drop(columns=['ASIN'], inplace=True)
    # Update "Blacklisted" column in MasterV
    master_v_df['Blacklisted'] = value
    # Save updated MasterV.csv
    master_v_df.to_csv(master_v_path, index=False)
    # Load Master_DB.csv into a DataFrame
    master_db_df = pd.read_csv(master_db_path)
    # Update "Blacklisted" column in Master_DB.csv
    master_db_df.loc[master_db_df['ASIN'] == asin, 'Blacklisted'] = value
    # Save updated Master_DB.csv
    master_db_df.to_csv(master_db_path, index=False)

def update_master_db_without_gogolescrappingpy(master_db_path, update_csv_path, manual_Update_Ouput):
    # Check if Master_DB.csv exists
    if os.path.exists(master_db_path):
        # Read Master_DB.csv
        master_db = pd.read_csv(master_db_path)
    else:
        # Create an empty DataFrame if Master_DB.csv doesn't exist
        master_db = pd.DataFrame(columns=['Title', 'Sales Rank: 30 days avg.', 'Sales Rank: 90 days avg.', 'Buy Box: Current',
                                           'Buy Box: Stock', 'New: Current', 'Referral Fee %', 'New, 3rd Party FBM: 30 days avg.',
                                           'Lowest FBM Seller', 'Count of retrieved live offers: New, FBM', 'ASIN',
                                           'Product Codes: EAN', 'Product Codes: UPC', 'Parent ASIN', 'Brand',
                                           'Prime Eligible (Amazon offer)', 'Custom_SKU', 'Google_Search_Link', 'Match',
                                           'Amazon_link', 'Extraction_Link', 'Store', 'Price', 'Min Price', 'Max Price',
                                           'Amazon_List_price', 'Can_List', 'Blacklisted', 'Curr_Listed?'])

    # Read Update_CSV.csv
    update_csv = pd.read_csv(update_csv_path)

    # Counter for added items
    added_count = 0

    # Iterate through rows in Update_CSV and update Master_DB accordingly
    for index, row in update_csv.iterrows():
        asin = row['ASIN']
        black_list = row.get('Blacklisted')
        ean = row.get('Product Codes: EAN')
        upc = row["Product Codes: UPC"]

        # Find the row in Master_DB that matches the ASIN
        master_row_index = master_db.index[master_db['ASIN'] == asin]

        if not master_row_index.empty:
            # ASIN found in Master_DB, update Blacklist if provided
            if black_list is not None:
                master_db.loc[master_row_index, 'Blacklisted'] = black_list
        else:
            # ASIN not found in Master_DB, add a new row
            new_row = pd.DataFrame({'Title': [row['Title']], 'Sales Rank: 30 days avg.': [row['Sales Rank: 30 days avg.']],
                                    'Sales Rank: 90 days avg.': [row['Sales Rank: 90 days avg.']], 'Buy Box: Current': [row['Buy Box: Current']],
                                    'Buy Box: Stock': [row['Buy Box: Stock']], 'New: Current': [row['New: Current']],
                                    'Referral Fee %': [row['Referral Fee %']], 'New, 3rd Party FBM: 30 days avg.': [row['New, 3rd Party FBM: 30 days avg.']],
                                    'Lowest FBM Seller': [row['Lowest FBM Seller']], 'Count of retrieved live offers: New, FBM': [row['Count of retrieved live offers: New, FBM']],
                                    'ASIN': [asin], 'Product Codes: EAN': [ean], 'Product Codes: UPC': [upc],
                                    'Parent ASIN': [row['Parent ASIN']], 'Brand': [row['Brand']],
                                    'Prime Eligible (Amazon offer)': [row['Prime Eligible (Amazon offer)']],
                                    'Custom_SKU': [row['Custom_SKU']], 'Google_Search_Link': [row['Google_Search_Link']],
                                    'Match': [row['Match']], 'Amazon_link': [row['Amazon_link']],
                                    'Extraction_Link': [row['Extraction_Link']], 'Store': [row['Store']],
                                    'Price': [row['Price']], 'Min Price': [row['Min Price']], 'Max Price': [row['Max Price']],
                                    'Amazon_List_price': [row['Amazon_List_price']], 'Can_List': [row['Can_List']],
                                    'Blacklisted': [black_list], 'Curr_Listed?': [row['Curr_Listed?']]})
            master_db = pd.concat([master_db, new_row], ignore_index=True)
            added_count += 1

    # Save the updated Master_DB
    master_db.to_csv(manual_Update_Ouput, index=False)

    print(f"{added_count} items were added to Master_DB successfully.")

def prepare_manual_only_import(input_file, output_file):
    # Read the input CSV file
    print("Working")
    spreadsheet = pd.read_csv(input_file)

    # Save temp data
    temp_blacklisted = spreadsheet['Black_list'].copy()
    temp_pid = spreadsheet['PID'].copy()

    # Delete the columns
    del spreadsheet['Black_list']
    del spreadsheet['PID']

    # Define a function to process each row
    def process_row(row, blacklisted, pid):
        upc = row['Product Codes: UPC']
        
        row['Prime Eligible (Amazon offer)'] = None
        # Custom_SKU
        row['Custom_SKU'] = "GB" + str(upc)

        # Google_Search_Link
        row['Google_Search_Link'] = create_google_shopping_link(upc, row['Title'])

        # Amazon_link
        row['Amazon_link'] = create_amazon_link(row['ASIN'])

        # Match
        row['Match'] = "Y"

        # Use temp PID data for Extraction_Link
        product_code = pid[row.name]  # Using row index to match
        if product_code == "none":
            row['Extraction_Link'] = None
        else:
            row['Extraction_Link'] = f'https://www.google.com/shopping/product/{product_code}/offers?q={upc}&prds=cid:{product_code},cond:1'
        
        # Set other fields to None or initial values
        row['Store'] = None
        row['Price'] = None
        row['Min Price'] = None
        row['Max Price'] = None
        row['Amazon_List_price'] = None
        row['Can_List'] = "N"
        row['Curr_Listed?'] = 0

        # Use temp Blacklisted data
        row['Blacklisted'] = blacklisted[row.name]  # Using row index to match

        return row

    # Apply the function to each row, passing the temp data as additional arguments
    spreadsheet = spreadsheet.apply(lambda row: process_row(row, temp_blacklisted, temp_pid), axis=1)

    # Define the columns order and rearrange the DataFrame
    columns_order = [
        'Title', 'Sales Rank: 30 days avg.', 'Sales Rank: 90 days avg.', 'Buy Box: Current',
        'Buy Box: Stock', 'New: Current', 'New: Highest', 'Referral Fee %',
        'New, 3rd Party FBM: 30 days avg.', 'Lowest FBM Seller', 
        'Count of retrieved live offers: New, FBM', 'ASIN', 'Product Codes: EAN', 
        'Product Codes: UPC', 'Parent ASIN', 'Brand', 'Prime Eligible (Amazon offer)',
        'Custom_SKU', 'Google_Search_Link', 'Match', 'Amazon_link', 'Extraction_Link', 
        'Store', 'Price', 'Min Price', 'Max Price', 'Amazon_List_price', 'Can_List', 'Blacklisted', 
        'Curr_Listed?'
    ]
    spreadsheet = spreadsheet.reindex(columns=columns_order, fill_value=None)

    # Save the prepared DataFrame to a new CSV file
    spreadsheet.to_csv(output_file, index=False)

    print(f"Prepared data saved to {output_file}")


import http.client
import mimetypes
from codecs import encode

base_url = "projectleapapi.agilensmartservices.com"
endpoint = "/uploadFile"
file_path = "C:\\Users\\Administrator\\Documents\\RA\\DataBaseFiles\\MasterV.csv"


def upload_file(file_path):
    conn = http.client.HTTPSConnection(base_url)
    dataList = []
    boundary = 'wL36Yn8afVp8Ag7AmP8qZ0SA4n1v9T'
    dataList.append(encode('--' + boundary))
    dataList.append(
        encode('Content-Disposition: form-data; name="csv_file"; filename="{0}"'.format(file_path)))

    fileType, _ = mimetypes.guess_type(file_path)
    fileType = fileType or 'application/octet-stream'
    dataList.append(encode('Content-Type: {}'.format(fileType)))
    dataList.append(encode(''))

    with open(file_path, 'rb') as f:
        dataList.append(f.read())
    dataList.append(encode('--' + boundary + '--'))
    dataList.append(encode(''))
    body = b'\r\n'.join(dataList)
    payload = body
    headers = {
        'accept': 'application/json',
        'Content-type': 'multipart/form-data; boundary={}'.format(boundary)
    }
    conn.request("POST", endpoint, payload, headers)
    res = conn.getresponse()
    data = res.read()
    print(data.decode("utf-8"))




def create_google_shopping_link(upc, title):
    base_url = "https://www.google.com/search?"
    query_parameters = {
        "q": f"{upc} {title}",
        "tbm": "shop"
    }
    encoded_query = urllib.parse.urlencode(query_parameters, quote_via=urllib.parse.quote_plus)
    return base_url + encoded_query

def create_amazon_link(asin):
    base_url = "https://www.amazon.com/dp/"
    return f"{base_url}{asin}?psc=1"

import csv
import os

def combine_csv_files(file1_path, file2_path, output_dir="KeepaExports"):
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Read the first row from one of the CSV files
    with open(file1_path, 'r', newline='', encoding='utf-8') as file1:
        reader = csv.reader(file1)
        first_row = next(reader)
    
    # Write the combined data to the output CSV file
    output_file = os.path.join(output_dir, "Keepa_Update.csv")
    with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)
        
        # Write the first row
        writer.writerow(first_row)
        
        # Append the remaining rows from all three files
        for file_path in [file1_path, file2_path]:
            with open(file_path, 'r', newline='', encoding='utf-8') as infile:
                next(infile)  # Skip the first row
                reader = csv.reader(infile)
                for row in reader:
                    writer.writerow(row)
    
    print("CSV files combined successfully. Output file:", output_file)

def update_master_v(keepa_file, master_file):
    # Step 1: Load Keepa_Update Data
    keepa_data = {}
    with open(keepa_file, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            asin = row['ASIN']
            # Store and clean Buy Box: Current, New: Current values for each ASIN, and store Lowest FBM Seller as is
            keepa_data[asin] = {
                'Buy Box: Current': clean_and_convert_price(row['Buy Box: Current']),
                'New: Current': clean_and_convert_price(row['New: Current']),
                'New: Highest': clean_and_convert_price(row['New: Highest']),
                'Lowest FBM Seller': row.get('Lowest FBM Seller')  # No cleaning needed
            }
    
    # Step 2: Read MasterV and Prepare Updated Data
    updated_rows = []
    with open(master_file, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        fieldnames = reader.fieldnames
        for row in reader:
            asin = row['ASIN']
            if asin in keepa_data:
                # Update values from Keepa_Update after cleaning and converting
                row['Buy Box: Current'] = keepa_data[asin]['Buy Box: Current']
                row['New: Current'] = keepa_data[asin]['New: Current']
                row['New: Highest'] = keepa_data[asin]['New: Highest']
                # Directly update Lowest FBM Seller
                row['Lowest FBM Seller'] = keepa_data[asin]['Lowest FBM Seller']
            updated_rows.append(row)
    
    # Step 3: Write Updates Back to MasterV
    with open(master_file, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)
    print("Done")

import pandas as pd

import pandas as pd

import pandas as pd

import pandas as pd

import pandas as pd

def update_master_v_UPC(master_file_path, update_file_path, output_file_path):
    """
    Updates EAN, UPC codes, Custom_SKU, and Extraction_Link in the master file based on an update file.
    
    Args:
    master_file_path (str): Path to the master CSV file.
    update_file_path (str): Path to the CSV file containing updates.
    output_file_path (str): Path to save the updated master CSV file.
    """
    # Load the data, ensuring all data is treated as string where necessary
    master_v = pd.read_csv(master_file_path, dtype={'Extraction_Link': str, 'Product Codes: UPC': str})
    master_v_upc_update = pd.read_csv(update_file_path, dtype={'Product Codes: UPC': str})
    
    # Loop through each row in the update DataFrame
    for index, update_row in master_v_upc_update.iterrows():
        try:
            # Find matching ASIN in MasterV
            mask = master_v['ASIN'] == update_row['ASIN']
            
            # Check if there are any matches before updating
            if mask.any():
                # Update the EAN and UPC codes
                master_v.loc[mask, 'Product Codes: EAN'] = update_row['Product Codes: EAN']
                master_v.loc[mask, 'Product Codes: UPC'] = update_row['Product Codes: UPC']
                
                # Update the Custom_SKU column
                master_v.loc[mask, 'Custom_SKU'] = "GB" + str(update_row['Product Codes: UPC'])
                
                # Update the Extraction_Link column
                for idx in master_v[mask].index:
                    if pd.notnull(master_v.at[idx, 'Extraction_Link']):
                        link_parts = master_v.at[idx, 'Extraction_Link'].split('cid:')
                        if len(link_parts) > 1:
                            pid = link_parts[1].split(',')[0]
                            upc = master_v.at[idx, 'Product Codes: UPC']
                            new_link = f'https://www.google.com/shopping/product/{pid}/offers?q={upc}&prds=cid:{pid},cond:1'
                            master_v.at[idx, 'Extraction_Link'] = new_link
                        else:
                            print(f"Error in row {idx}: Unexpected link format")
        except Exception as e:
            print(f"Error in row {index}: {str(e)}")
            continue  # Skip to the next row

    # Save the updated DataFrame back to CSV
    master_v.to_csv(output_file_path, index=False)
    print(f"Update complete! The file {output_file_path} has been saved.")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell  # Import to handle MergedCell objects

def update_excel_from_csv(master_csv_path, excel_path):
    # Load data from CSV with low_memory set to False to avoid DtypeWarnings
    master_data = pd.read_csv(master_csv_path, low_memory=False)
    print("CSV loaded successfully. Number of rows read:", len(master_data))

    # Load the Excel workbook and specify which sheet to use
    wb = load_workbook(excel_path)
    ws = wb.active
    print("Excel workbook loaded successfully.")

    # Handle headers on the third row, not the second
    header_row = ws[3]
    column_mapping = {
        'Product Codes: UPC': 'External Product ID',
        'Custom_SKU': 'Contribution SKU',
        'ASIN': 'Merchant Suggested ASIN',
        'Amazon_List_price': 'Your Price USD (US)',
        'Curr_Listed?': 'Quantity (US)'
    }

    # Creating a dictionary from header cells, handling merged cells
    header_dict = {}
    for cell in header_row:
        print(f"Reading header: {cell.value} at {cell.coordinate}")
        if cell.value is not None:
            if isinstance(cell, MergedCell):  # If the cell is a merged cell
                actual_cell = ws[cell.coordinate]
                header_dict[cell.value] = actual_cell.column_letter
            else:
                header_dict[cell.value] = cell.column_letter

    print("Header mapping:", header_dict)

    # Clear existing data from relevant columns in Excel starting from the fourth row
    for header in column_mapping.values():
        if header in header_dict:
            col_letter = header_dict[header]
            for row in range(6, ws.max_row + 1):
                ws[f'{col_letter}{row}'].value = None

    # Fill new data
    for index, row in master_data.iterrows():
        for master_col, excel_col in column_mapping.items():
            if excel_col in header_dict:
                col_letter = header_dict[excel_col]
                # Copy data to the correct column in Excel, starting from the fourth row
                ws[f'{col_letter}{index + 6}'].value = row[master_col]

    # Save the workbook
    wb.save(excel_path)
    print("Excel file has been updated and saved successfully!")



def remove_numeric_values(filename, columns):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(filename)
    
    # Remove numeric values from specified columns
    for column in columns:
        df[column] = df[column].apply(lambda x: '' if str(x).replace('.', '').isdigit() else x)
    
    # Save the modified DataFrame back to a CSV file
    df.to_csv(filename, index=False)


def combine_and_filter_exports(input_files, master_file_path, blacklist_file_path, output_file):
    # Initialize a DataFrame to hold all combined data
    all_data = pd.DataFrame()
    
    # Attempt to read and combine all input files
    total_initial_rows = 0
    for file_path in input_files:
        try:
            if os.path.exists(file_path):
                data = pd.read_csv(file_path)
                all_data = pd.concat([all_data, data], ignore_index=True)
                total_initial_rows += len(data)
            else:
                print(f"File '{file_path}' not found.")
        except Exception as e:
            print(f"Failed to read {file_path}: {e}")

    # Drop duplicates based on 'ASIN'
    all_data.drop_duplicates(subset='ASIN', inplace=True)
    
    try:
        # Load the master file
        if os.path.exists(master_file_path):
            master_data = pd.read_csv(master_file_path)
        else:
            print(f"Master file '{master_file_path}' not found.")
            master_data = pd.DataFrame()

        # Load the blacklist file
        if os.path.exists(blacklist_file_path):
            blacklist_data = pd.read_csv(blacklist_file_path)
        else:
            print(f"Blacklist file '{blacklist_file_path}' not found.")
            blacklist_data = pd.DataFrame()

    except Exception as e:
        print(f"Failed to load verification or blacklist data: {e}")
        return

    # Filter out ASINs that are in the master list or the blacklist
    filtered_data = all_data[
        ~all_data['ASIN'].isin(master_data.get('ASIN', [])) &
        ~all_data['ASIN'].isin(blacklist_data.get('ASIN', []))
    ]

    # Filter to include only ASINs that appear in all input files
    asin_counts = all_data['ASIN'].value_counts()
    asins_in_all_files = asin_counts[asin_counts == len(input_files)].index
    final_data = filtered_data[filtered_data['ASIN'].isin(asins_in_all_files)]

    # Output how many rows were in the input and how many are in the final output
    final_row_count = len(final_data)
    print(f"Total rows checked from input files: {total_initial_rows}")
    print(f"Rows in the final output: {final_row_count}")

    # Write the filtered data to the output file
    try:
        final_data.to_csv(output_file, index=False)
        print(f"Data exported successfully to {output_file}")
    except Exception as e:
        print(f"Failed to write to {output_file}: {e}")
